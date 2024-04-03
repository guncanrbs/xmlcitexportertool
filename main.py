import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from tkinter import filedialog
from tkinter import Tk, Button, Label, Frame, Text, messagebox

def parse_xml(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()

    data_dict = {}
    declaration = root.find('.//Declaration[@Id="DEC"]')

    if declaration is not None:
        dokuienv5 = declaration.find('.//DokUIENv5')
        if dokuienv5 is not None:
            taks_menesis = dokuienv5.find('.//TaksMenesis')
            if taks_menesis is not None:
                for r_number in range(1, 33):  # Loop through R01 to R32
                    r_element = dokuienv5.find(f'.//R{r_number:02d}')
                    if r_element is not None:
                        data_dict[f'R{r_number:02d}'] = float(r_element.text) if r_element.text else None

                return taks_menesis.text, data_dict

    return None, None

def write_to_excel(excel_file, data_dict, column):
    try:
        wb = load_workbook(excel_file)
        ws = wb.active

        # Write header rows
        ws['A1'] = 'R Row'
        ws[f'{column}1'] = 'Data'

        # Write months in header row
        for month in range(1, 13):
            ws[f'{get_column_letter(month + 1)}1'] = f'{get_month_name(month)}'

        # Write data to Excel in corresponding columns
        for idx, (tag, value) in enumerate(data_dict.items(), start=2):
            ws[f'A{idx}'] = tag
            ws[f'{column}{idx}'] = value

        # Set data column as Number format
        for cell in ws[f'{column}']:
            cell.number_format = '0'

        wb.save(excel_file)
    except Exception as e:
        print(f"Error writing data to Excel: {e}")

def get_month_name(month):
    month_names = [
        'January', 'February', 'March', 'April', 'May', 'June',
        'July', 'August', 'September', 'October', 'November', 'December'
    ]
    return month_names[month - 1]

def main():
    Tk().withdraw()  # Hide the main window
    file_paths = filedialog.askopenfilenames(title="Select XML files", filetypes=[("XML files", "*.xml")])

    if not file_paths:
        print("No files selected. Exiting.")
        return

    excel_file_path = r'ExcelResult\OutputExcel.xlsx'
    for file_path in file_paths:
        taks_menesis, data_dict = parse_xml(file_path)

        if taks_menesis is not None and data_dict is not None:
            print(f"Data extracted from {file_path} with TaksMenesis: {taks_menesis}")
            print(data_dict)

            column = get_column_letter(int(taks_menesis) + 1)  # Convert TaksMenesis to corresponding column
            write_to_excel(excel_file_path, data_dict, column)

    print(f"Data from {len(file_paths)} XML files has been successfully written to {excel_file_path}")

    messagebox.showinfo("Extraction Complete",
                        "Data from XML files has been successfully written to Excel. You can now exit.")


def show_instructions():
    instructions_popup = Tk()
    instructions_popup.title("Instructions")

    instructions_text = """1. Click 'Open XML Files' to select XML files - CIT declarations - use only declarations from EDS system.
2. The data will be extracted and written to the Excel file - XMLExporterTool\ExcelResult
3. Each file's data will be written in a separate column based on the declaration month.
4. Be sure to have an empty Excel file before clicking 'Open XML files', it can be found in XMLExporterTool\CLEAN EXCEL
5. After each time you export the data to Excel, copy CLEAN Excel file to Excel Result folder file.
6. You are ready to go!'
"""

    text_widget = Text(instructions_popup, wrap="word", height=10, width=50)
    text_widget.insert("1.0", instructions_text)
    text_widget.pack(padx=10, pady=10)

    close_button = Button(instructions_popup, text="Close", command=instructions_popup.destroy)
    close_button.pack(pady=10)


# Create the main window
root = Tk()
root.title("CIT Exporter XML")

# Design for the label
label_frame = Frame(root, bg="#444", padx=10, pady=10)
label_frame.pack(padx=20, pady=20)

label = Label(label_frame, text="XML CIT Exporter", font=("Arial Narrow", 16), fg="white", bg="#444")
label.pack()

# Design for the button
button_frame = Frame(root)
button_frame.pack(padx=20, pady=20)

# Correct placement of the command attribute
button = Button(button_frame, text="Open XML Files", command=main, font=("Arial Narrow", 12), bg="gray", fg="white")
button.pack()

# Instructions button
instructions_button = Button(button_frame, text="Instructions", command=show_instructions, font=("Arial Narrow", 12), bg="gray", fg="white")
instructions_button.pack()

# Run the Tkinter main loop
root.mainloop()

if __name__ == "__main__":
    main()